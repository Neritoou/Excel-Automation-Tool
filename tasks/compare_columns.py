"""
Tarea: Comparar columnas entre dos archivos Excel.

Compara cada valor de la columna seleccionada del Excel 1 contra
todos los valores de la columna seleccionada del Excel 2.
- Si el valor existe en ambos → contorno VERDE en columna+1
- Si el valor NO existe        → contorno ROJO  en columna+1
"""

from typing import Any
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet
from core.base_task import BaseTask, TaskParam, TaskResult
from core.excel_handler import ExcelHandler


class CompareColumnsTask(BaseTask):

    task_id = "compare_columns"
    task_name = "Comparar Columnas"
    task_description = (
        "Compara los datos de una columna entre dos archivos Excel. "
        "Marca en verde las coincidencias y en rojo las diferencias."
    )
    task_icon = "🔍"

    GREEN_BORDER = Border(
        left=Side(style="medium", color="228B22"),
        right=Side(style="medium", color="228B22"),
        top=Side(style="medium", color="228B22"),
        bottom=Side(style="medium", color="228B22"),
    )
    RED_BORDER = Border(
        left=Side(style="medium", color="DC143C"),
        right=Side(style="medium", color="DC143C"),
        top=Side(style="medium", color="DC143C"),
        bottom=Side(style="medium", color="DC143C"),
    )
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    def get_params(self) -> list[TaskParam]:
        return [
            TaskParam("file_1", "Archivo Excel 1", "file", group="Excel 1"),
            TaskParam("sheet_1", "Hoja", "sheet", depends_on="file_1", group="Excel 1"),
            TaskParam("column_1", "Columna a comparar", "column", depends_on="sheet_1", group="Excel 1"),
            TaskParam("file_2", "Archivo Excel 2", "file", group="Excel 2"),
            TaskParam("sheet_2", "Hoja", "sheet", depends_on="file_2", group="Excel 2"),
            TaskParam("column_2", "Columna a comparar", "column", depends_on="sheet_2", group="Excel 2"),
        ]

    def validate(self, params: dict[str, Any]) -> tuple[bool, str]:
        for key in ("file_1", "sheet_1", "column_1", "file_2", "sheet_2", "column_2"):
            if not params.get(key):
                return False, f"Falta el parámetro: {key}"
        return True, ""

    def _run(self, params: dict[str, Any]) -> TaskResult:
        handler = ExcelHandler()

        wb1 = handler.load(params["file_1"])
        wb2 = handler.load(params["file_2"])

        ws1: Worksheet = wb1[str(params["sheet_1"])]  # type: ignore[assignment]
        ws2: Worksheet = wb2[str(params["sheet_2"])]  # type: ignore[assignment]

        col1_idx: int = int(params["column_1"])
        col2_idx: int = int(params["column_2"])

        vals1 = handler.read_column_values(ws1, col1_idx)
        vals2 = handler.read_column_values(ws2, col2_idx)

        set2: set[str] = {self._normalize(v) for _, v in vals2}
        set1: set[str] = {self._normalize(v) for _, v in vals1}

        mark_col_1: int = col1_idx + 1
        matches_1, diffs_1 = self._mark_cells(ws1, vals1, set2, mark_col_1)

        mark_col_2: int = col2_idx + 1
        matches_2, diffs_2 = self._mark_cells(ws2, vals2, set1, mark_col_2)

        header_font = Font(bold=True, size=10)
        ws1.cell(row=1, column=mark_col_1, value="Resultado").font = header_font  # type: ignore[union-attr]
        ws2.cell(row=1, column=mark_col_2, value="Resultado").font = header_font  # type: ignore[union-attr]

        file_1: str = str(params["file_1"])
        file_2: str = str(params["file_2"])
        wb1.save(file_1)
        wb2.save(file_2)

        return TaskResult(
            success=True,
            message=(
                f"Comparación completada.\n"
                f"  Excel 1 → {matches_1} coincidencias, {diffs_1} diferencias\n"
                f"  Excel 2 → {matches_2} coincidencias, {diffs_2} diferencias"
            ),
            output_files=[file_1, file_2],
            details={
                "excel_1": {"matches": matches_1, "diffs": diffs_1},
                "excel_2": {"matches": matches_2, "diffs": diffs_2},
            },
        )

    def _mark_cells(
        self, ws: Worksheet, values: list[tuple[int, Any]], reference_set: set[str], mark_col: int
    ) -> tuple[int, int]:
        matches: int = 0
        diffs: int = 0
        for row, val in values:
            cell = ws.cell(row=row, column=mark_col)
            if self._normalize(val) in reference_set:
                cell.value = "✔"  # type: ignore[union-attr]
                cell.border = self.GREEN_BORDER  # type: ignore[union-attr]
                cell.fill = self.GREEN_FILL  # type: ignore[union-attr]
                cell.font = Font(color="006100", bold=True)  # type: ignore[union-attr]
                matches += 1
            else:
                cell.value = "✘"  # type: ignore[union-attr]
                cell.border = self.RED_BORDER  # type: ignore[union-attr]
                cell.fill = self.RED_FILL  # type: ignore[union-attr]
                cell.font = Font(color="9C0006", bold=True)  # type: ignore[union-attr]
                diffs += 1
        return matches, diffs

    @staticmethod
    def _normalize(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()
