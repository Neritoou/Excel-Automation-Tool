"""
Tarea: VLOOKUP entre dos archivos Excel.

Simula la función BUSCARV de Excel:
  1. Seleccionar columna clave en Excel 1 (destino)
  2. Seleccionar columna clave en Excel 2 (origen)
  3. Seleccionar columna de datos en Excel 2 (valor a traer)
  4. Para cada fila del Excel 1, busca la clave en Excel 2
     y copia el valor correspondiente al Excel 1 en columna+1

Resultado: los datos del Excel 2 aparecen en el Excel 1
junto a cada coincidencia. Las no encontradas se marcan "N/A".
"""

from typing import Any
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from core.base_task import BaseTask, TaskParam, TaskResult
from core.excel_handler import ExcelHandler


class VlookupTask(BaseTask):

    task_id = "vlookup"
    task_name = "BUSCARV (VLOOKUP)"
    task_description = (
        "Busca valores de una columna clave del Excel 1 en el Excel 2, "
        "y trae los datos de una columna adicional del Excel 2 al Excel 1. "
        "Similar a la función BUSCARV de Excel."
    )
    task_icon = "🔗"

    FOUND_FILL = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    NOT_FOUND_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    FOUND_BORDER = Border(
        left=Side(style="thin", color="3B82F6"), right=Side(style="thin", color="3B82F6"),
        top=Side(style="thin", color="3B82F6"), bottom=Side(style="thin", color="3B82F6"),
    )

    def get_params(self) -> list[TaskParam]:
        return [
            TaskParam("file_dest", "Excel destino", "file", group="Excel 1 (destino)"),
            TaskParam("sheet_dest", "Hoja", "sheet", depends_on="file_dest", group="Excel 1 (destino)"),
            TaskParam("col_key_dest", "Columna clave", "column", depends_on="sheet_dest", group="Excel 1 (destino)"),
            TaskParam("file_src", "Excel origen", "file", group="Excel 2 (origen)"),
            TaskParam("sheet_src", "Hoja", "sheet", depends_on="file_src", group="Excel 2 (origen)"),
            TaskParam("col_key_src", "Columna clave", "column", depends_on="sheet_src", group="Excel 2 (origen)"),
            TaskParam("col_value_src", "Columna de datos a traer", "column", depends_on="sheet_src", group="Excel 2 (origen)"),
        ]

    def validate(self, params: dict[str, Any]) -> tuple[bool, str]:
        required: tuple[str, ...] = (
            "file_dest", "sheet_dest", "col_key_dest",
            "file_src", "sheet_src", "col_key_src", "col_value_src",
        )
        for key in required:
            if not params.get(key):
                return False, f"Falta: {key}"
        return True, ""

    def _run(self, params: dict[str, Any]) -> TaskResult:
        handler = ExcelHandler()

        file_src: str = str(params["file_src"])
        wb_src = handler.load(file_src)
        ws_src: Worksheet = wb_src[str(params["sheet_src"])]  # type: ignore[assignment]

        col_key_src: int = int(params["col_key_src"])
        col_value_src: int = int(params["col_value_src"])

        keys_src = handler.read_column_values(ws_src, col_key_src)
        vals_src = handler.read_column_values(ws_src, col_value_src)

        row_to_val: dict[int, Any] = {row: val for row, val in vals_src}
        lookup: dict[str, Any] = {}
        for row, key in keys_src:
            norm: str = self._normalize(key)
            if norm not in lookup and row in row_to_val:
                lookup[norm] = row_to_val[row]

        src_header: str = self._get_header(ws_src, col_value_src)
        wb_src.close()

        file_dest: str = str(params["file_dest"])
        wb_dest = handler.load(file_dest)
        ws_dest: Worksheet = wb_dest[str(params["sheet_dest"])]  # type: ignore[assignment]

        col_key_dest: int = int(params["col_key_dest"])
        result_col: int = col_key_dest + 1

        header_cell = ws_dest.cell(row=1, column=result_col, value=f"VLOOKUP: {src_header}")
        header_cell.font = Font(bold=True, size=10)  # type: ignore[union-attr]

        keys_dest = handler.read_column_values(ws_dest, col_key_dest)
        found: int = 0
        not_found: int = 0

        for row, key in keys_dest:
            norm = self._normalize(key)
            cell = ws_dest.cell(row=row, column=result_col)

            if norm in lookup:
                cell.value = lookup[norm]  # type: ignore[union-attr]
                cell.fill = self.FOUND_FILL  # type: ignore[union-attr]
                cell.border = self.FOUND_BORDER  # type: ignore[union-attr]
                cell.font = Font(color="1E40AF", size=10)  # type: ignore[union-attr]
                found += 1
            else:
                cell.value = "N/A"  # type: ignore[union-attr]
                cell.fill = self.NOT_FOUND_FILL  # type: ignore[union-attr]
                cell.font = Font(color="92400E", italic=True, size=10)  # type: ignore[union-attr]
                not_found += 1

        wb_dest.save(file_dest)

        return TaskResult(
            success=True,
            message=(
                f"VLOOKUP completado: {found} valores encontrados, "
                f"{not_found} no encontrados (N/A)"
            ),
            output_files=[file_dest],
            details={"found": found, "not_found": not_found},
        )

    @staticmethod
    def _normalize(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip().lower()

    @staticmethod
    def _get_header(ws: Worksheet, col_idx: int) -> str:
        val: Any = ws.cell(row=1, column=col_idx).value
        return str(val) if val is not None else f"Col {col_idx}"
