"""
Tarea: Eliminar / Marcar Duplicados.

Analiza una columna y permite:
  - Marcar duplicados con color (sin eliminar)
  - Eliminar filas duplicadas conservando la primera aparición

Resultado en columna+1: "ÚNICO" (verde) o "DUPLICADO" (rojo).
"""

from typing import Any
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet
from core.base_task import BaseTask, TaskParam, TaskResult
from core.excel_handler import ExcelHandler


class RemoveDuplicatesTask(BaseTask):

    task_id = "remove_duplicates"
    task_name = "Duplicados"
    task_description = (
        "Detecta valores duplicados en una columna. Puede marcarlos con "
        "color o eliminar las filas duplicadas conservando la primera aparición."
    )
    task_icon = "🧹"

    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GREEN_BORDER = Border(
        left=Side(style="thin", color="228B22"), right=Side(style="thin", color="228B22"),
        top=Side(style="thin", color="228B22"), bottom=Side(style="thin", color="228B22"),
    )
    RED_BORDER = Border(
        left=Side(style="thin", color="DC143C"), right=Side(style="thin", color="DC143C"),
        top=Side(style="thin", color="DC143C"), bottom=Side(style="thin", color="DC143C"),
    )

    def get_params(self) -> list[TaskParam]:
        return [
            TaskParam("file", "Archivo Excel", "file", group="Archivo"),
            TaskParam("sheet", "Hoja", "sheet", depends_on="file", group="Archivo"),
            TaskParam("column", "Columna clave", "column", depends_on="sheet", group="Archivo"),
            TaskParam("delete", "¿Eliminar filas duplicadas?", "bool", default=False, group="Opciones"),
        ]

    def validate(self, params: dict[str, Any]) -> tuple[bool, str]:
        for key in ("file", "sheet", "column"):
            if not params.get(key):
                return False, f"Falta: {key}"
        return True, ""

    def _run(self, params: dict[str, Any]) -> TaskResult:
        handler = ExcelHandler()
        file_path: str = str(params["file"])
        wb = handler.load(file_path)
        ws: Worksheet = wb[str(params["sheet"])]  # type: ignore[assignment]
        col_idx: int = int(params["column"])
        delete_rows: bool = bool(params.get("delete", False))
        mark_col: int = col_idx + 1

        ws.cell(row=1, column=mark_col, value="Estado").font = Font(bold=True, size=10)  # type: ignore[union-attr]

        values = handler.read_column_values(ws, col_idx)
        seen: dict[str, int] = {}
        duplicates: list[int] = []
        uniques: int = 0

        for row, val in values:
            norm: str = str(val).strip().lower() if val is not None else ""
            cell = ws.cell(row=row, column=mark_col)

            if norm in seen:
                cell.value = "DUPLICADO"  # type: ignore[union-attr]
                cell.fill = self.RED_FILL  # type: ignore[union-attr]
                cell.border = self.RED_BORDER  # type: ignore[union-attr]
                cell.font = Font(color="9C0006", bold=True, size=9)  # type: ignore[union-attr]
                duplicates.append(row)
            else:
                seen[norm] = row
                cell.value = "ÚNICO"  # type: ignore[union-attr]
                cell.fill = self.GREEN_FILL  # type: ignore[union-attr]
                cell.border = self.GREEN_BORDER  # type: ignore[union-attr]
                cell.font = Font(color="006100", bold=True, size=9)  # type: ignore[union-attr]
                uniques += 1

        deleted: int = 0
        if delete_rows and duplicates:
            for row in sorted(duplicates, reverse=True):
                ws.delete_rows(row)
                deleted += 1

        wb.save(file_path)

        action = f"y {deleted} filas eliminadas" if deleted else "(solo marcados, sin eliminar)"
        return TaskResult(
            success=True,
            message=(
                f"Análisis completado: {uniques} únicos, "
                f"{len(duplicates)} duplicados {action}"
            ),
            output_files=[file_path],
            details={"uniques": uniques, "duplicates": len(duplicates), "deleted": deleted},
        )
