from dataclasses import dataclass
from typing import Any, NamedTuple

import xlwings as xw
from openpyxl import load_workbook


# --- TYPES ---

RGB = tuple[int, int, int]

@dataclass(frozen=True, slots=True)
class CellStyle:
    """ Descriptor inmutable de formato para una celda Excel."""
    fill: RGB | None = None
    font_color: RGB | None = None
    bold: bool = False
    italic: bool = False
    font_size: int | None = None

class CellValue(NamedTuple):
    """Par (fila, valor) de una celda leída."""

    row: int
    value: Any

# --- HANDLER ---

class ExcelHandler:
    """Utilidades estáticas para lectura/escritura de archivos Excel."""

    @staticmethod
    def get_sheet_names(path: str) -> list[str]:
        """Devuelve los nombres de las hojas sin abrir Excel."""
        wb = load_workbook(path, read_only=True)
        names = wb.sheetnames
        wb.close()
        return names

    @staticmethod
    def get_column_headers(path: str, sheet_name: str) -> list[str]:
        """Devuelve los encabezados (fila 1) de la hoja indicada."""
        wb = load_workbook(path, read_only=True)
        ws = wb[sheet_name]
        headers: list[str] = [
            str(cell.value) if cell.value is not None else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        wb.close()
        return headers

    @staticmethod
    def load(path: str) -> xw.Book:
        """
        Abre el workbook vía xlwings.

        Si el archivo ya está abierto en Excel, se conecta a esa instancia.
        """
        return xw.Book(path)

    @staticmethod
    def read_column_values(ws: xw.Sheet, col_idx: int, start_row: int = 2) -> list[CellValue]:
        """
        Lee los valores de una columna desde start_row.

        Args:
            ws: Hoja de xlwings.
            col_idx: Índice de columna (1-based).
            start_row: Fila inicial (default 2, omite header).
        """
        last_row = ws.used_range.last_cell.row
        if last_row < start_row:
            return []

        raw = ws.range((start_row, col_idx), (last_row, col_idx)).value

        if not isinstance(raw, list):
            raw = [raw]

        return [CellValue(start_row + i, val) for i, val in enumerate(raw) if val is not None]

    @staticmethod
    def set_cell(ws: xw.Sheet, row: int, col: int, value: Any = None, style: CellStyle | None = None) -> None:
        """
        Escribe un valor y aplica formato a una celda.

        Args:
            ws: Hoja de xlwings.
            row: Fila (1-based).
            col: Columna (1-based).
            value: Valor a escribir (None = no modificar).
            style: Descriptor de formato (None = sin formato).
        """
        cell = ws.range((row, col))

        if value is not None:
            cell.value = value

        if style is None:
            return

        if style.fill is not None:
            cell.color = style.fill

        cell.font.bold = style.bold
        cell.font.italic = style.italic

        if style.font_color is not None:
            cell.font.color = style.font_color

        if style.font_size is not None:
            cell.font.size = style.font_size

    @staticmethod
    def color_rows(ws: xw.Sheet, col: int, rows: list[int], fill: RGB) -> None:
        """
        Aplica color de relleno a múltiples filas de una columna en una sola operación COM.

        Mucho más rápido que llamar set_cell en un loop — construye un rango
        discontinuo (Union) y aplica el color de golpe.

        Args:
            ws: Hoja de xlwings.
            col: Columna (1-based).
            rows: Lista de filas (1-based) a colorear.
            fill: Color como (R, G, B).
        """
        if not rows:
            return

        api = ws.api
        rng = api.Cells(rows[0], col)

        for r in rows[1:]:
            rng = api.Application.Union(rng, api.Cells(r, col))

        rng.Interior.Color = fill[2] * 65536 + fill[1] * 256 + fill[0]